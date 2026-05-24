"""Generate downloadable Excel templates from TemplateSpec.

The /api/admin/upload-template/<bronze_table> endpoint streams the
result of generate_template_workbook() to the user. We don't ship
pre-built .xlsx files in the repo — they'd drift from the schema and
inflate the diff every time a column changes.

Layout (per template):

  Sheet "<sheet_name>" (e.g. "pc_policy")
    Row 1: column names — bold, light-blue fill, matches Bronze DDL exactly
    Row 2: helper hint — italic, muted, "Type: <logical> · <required?>"
    Row 3: example values — gray italic, pre-filled to show shape
    Row 4: empty (start typing here)
    Row N+: user-provided data

  Sheet "_README"
    Hand-friendly instructions: what to fill in, what NOT to change,
    where to send questions.

Why two header rows: the conversion step skips rows 1+2+3 by default,
so the user can leave the example row in or delete it without breaking
the parse. The README sheet is detected by name and ignored on read.
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from packages.uploads.schemas import TemplateSpec, get_template

_HEADER_FILL = PatternFill(start_color="DDE8F4", end_color="DDE8F4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="0F172A")
_HINT_FONT = Font(italic=True, color="64748B", size=10)
_EXAMPLE_FONT = Font(italic=True, color="94A3B8")
_README_TITLE_FONT = Font(bold=True, size=14)
_README_BODY_FONT = Font(size=11)

# Excel column widths — picked to fit longer names without auto-fit faff.
_DEFAULT_COL_WIDTH = 22


def _hint_for(col_logical_type: str, required: bool) -> str:
    req = "REQUIRED" if required else "optional"
    pretty = {
        "string": "text",
        "integer": "whole number",
        "decimal": "decimal number",
        "date": "date (YYYY-MM-DD)",
        "datetime": "datetime (YYYY-MM-DD HH:MM:SS)",
        "boolean": "true / false",
    }.get(col_logical_type, col_logical_type)
    return f"{pretty} · {req}"


def generate_template_workbook(bronze_table: str) -> bytes:
    """Build the .xlsx for `bronze_table` and return its bytes.

    Raises KeyError if the template isn't registered."""
    spec = get_template(bronze_table)
    if spec is None:
        raise KeyError(f"No upload template registered for bronze_table={bronze_table!r}")

    wb = openpyxl.Workbook()
    # Default sheet → data sheet
    ws = wb.active
    ws.title = spec.sheet_name

    # Row 1: column names
    for col_idx, col in enumerate(spec.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = _DEFAULT_COL_WIDTH

    # Row 2: type + required hint
    for col_idx, col in enumerate(spec.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=_hint_for(col.logical_type, col.required))
        cell.font = _HINT_FONT

    # Row 3: example values
    for col_idx, col in enumerate(spec.columns, start=1):
        if col.example is not None:
            cell = ws.cell(row=3, column=col_idx, value=col.example)
            cell.font = _EXAMPLE_FONT

    ws.freeze_panes = "A4"  # keep header + hint + example visible while scrolling

    # README sheet — instructions
    readme = wb.create_sheet("_README")
    readme.column_dimensions["A"].width = 90
    readme.cell(row=1, column=1, value=f"RegulAI upload template — {spec.bronze_table}").font = _README_TITLE_FONT
    readme.cell(row=3, column=1, value=spec.description).font = _README_BODY_FONT
    readme.cell(row=5, column=1, value="HOW TO USE THIS FILE").font = _README_TITLE_FONT
    instructions = [
        "1. Open the data sheet (next tab to the left).",
        "2. Row 1 is the column header — DO NOT rename or reorder columns.",
        "3. Row 2 explains each column's expected type and whether it's required.",
        "4. Row 3 is an example. Keep it, delete it, or overwrite it — the upload skips the first 3 rows.",
        "5. Add your data starting at row 4. One policy per row.",
        "6. Save the file and upload it at /admin/upload in the RegulAI app.",
        "",
        "WHAT THE SYSTEM DOES NEXT",
        "• Validates your column headers match the template exactly.",
        "• Converts the data to Parquet (RegulAI's internal storage).",
        "• Loads it into BRONZE.GW_PC_POLICY in Snowflake.",
        "• You then click 'Process' to run the medallion pipeline (Bronze → Silver → Gold → Validate).",
        "",
        "TROUBLESHOOTING",
        "• If the upload fails with 'column mismatch', re-download the template — column names changed.",
        "• If a cell's type is wrong (e.g. text in an integer column), the upload reports the row + column.",
        "• Empty cells in optional columns are fine; empty cells in REQUIRED columns are rejected.",
    ]
    for i, line in enumerate(instructions):
        readme.cell(row=7 + i, column=1, value=line).font = _README_BODY_FONT

    # Serialize to bytes
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_filename_for(bronze_table: str) -> str:
    spec = get_template(bronze_table)
    if spec is None:
        raise KeyError(f"No upload template registered for bronze_table={bronze_table!r}")
    return spec.template_filename

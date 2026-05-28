"""Convert an uploaded .xlsx into a typed Parquet file.

Called by:
  - The /api/admin/upload endpoint immediately after save (so we fail
    fast on schema mismatch before involving Dagster)
  - The Dagster op as defense in depth (recovers if the file is on
    disk but the .parquet got cleaned up)

Parse contract:
  - The data sheet name MUST match TemplateSpec.sheet_name
  - Row 1 = column headers; columns must match the template exactly
    (same names, same order is not required, but every required column
    must be present)
  - Rows 2 and 3 are skipped (hint + example rows from the template)
  - Data starts at row 4
  - Empty trailing rows are tolerated
  - Empty cells in REQUIRED columns raise ConversionError with row/col

The conversion is intentionally strict — we'd rather fail loud and tell
the user what's wrong than ingest malformed rows that blow up Silver/Gold.
"""

import datetime as dt
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl.cell.cell import Cell

from packages.uploads.schemas import (
    ColumnSpec,
    LogicalType,
    TemplateSpec,
    get_template,
    pyarrow_type_for,
)
from packages.uploads.storage import (
    UploadRecord,
    parquet_path_for,
    upload_dir,
)


class ConversionError(ValueError):
    """The uploaded file doesn't match the template, or has bad data.
    Caller (the FastAPI route or the Dagster op) should surface this
    to the user verbatim — the messages already say what to fix."""


@dataclass(frozen=True)
class ConversionResult:
    parquet_path: Path
    row_count: int
    skipped_cells: int  # rows with optional cells left empty (for telemetry)


# ── Cell type coercion ─────────────────────────────────────────────


def _coerce_cell(value, col: ColumnSpec, row_num: int) -> object | None:
    """Coerce one Excel cell to the logical type from the template.
    Returns None on empty cells (validation happens upstream)."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None

    lt: LogicalType = col.logical_type

    try:
        if lt == "string":
            return str(value)
        if lt == "integer":
            if isinstance(value, bool):
                # openpyxl sees TRUE/FALSE as bools; reject in integer columns
                raise ConversionError(
                    f"Row {row_num}, column {col.name!r}: expected integer, got boolean."
                )
            return int(value)
        if lt == "decimal":
            # PyArrow Decimal128 accepts Python Decimal; let it cast from float
            from decimal import Decimal
            return Decimal(str(value))
        if lt == "date":
            if isinstance(value, dt.datetime):
                return value.date()
            if isinstance(value, dt.date):
                return value
            return dt.date.fromisoformat(str(value)[:10])
        if lt == "datetime":
            if isinstance(value, dt.datetime):
                return value
            if isinstance(value, dt.date):
                return dt.datetime.combine(value, dt.time())
            # Try ISO formats with and without time
            s = str(value).strip()
            try:
                return dt.datetime.fromisoformat(s)
            except ValueError:
                return dt.datetime.fromisoformat(s + "T00:00:00")
        if lt == "boolean":
            if isinstance(value, bool):
                return value
            s = str(value).strip().lower()
            if s in ("true", "yes", "y", "1"):
                return True
            if s in ("false", "no", "n", "0"):
                return False
            raise ConversionError(
                f"Row {row_num}, column {col.name!r}: cannot parse {value!r} as boolean."
            )
    except ConversionError:
        raise
    except Exception as e:
        raise ConversionError(
            f"Row {row_num}, column {col.name!r}: cannot parse {value!r} as {lt} ({e})."
        ) from e

    raise ConversionError(f"Unsupported logical_type {lt} for column {col.name!r}.")


# ── Top-level conversion ─────────────────────────────────────────────


def convert_uploaded_xlsx_to_parquet(record: UploadRecord) -> ConversionResult:
    """Convert the file referenced by `record` into a typed Parquet.

    Writes to: materialized/uploads/<upload_id>/parquet/<bronze_table>.parquet.

    Raises ConversionError with an actionable message on any schema or
    data problem. Does NOT update the UploadRecord status — the caller
    does that based on the outcome.
    """
    spec = get_template(record.bronze_table)
    if spec is None:
        raise ConversionError(
            f"No upload template registered for bronze_table={record.bronze_table!r}. "
            f"Add a TemplateSpec to packages/uploads/schemas.py first."
        )

    xlsx_path = upload_dir(record.upload_id) / "original" / record.filename
    if not xlsx_path.exists():
        raise ConversionError(
            f"Uploaded file not found on disk: {xlsx_path}. "
            f"This usually means the upload directory was cleaned up."
        )

    try:
        wb = openpyxl.load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        raise ConversionError(
            f"Cannot open {record.filename!r} as a workbook ({e}). "
            f"Re-save as .xlsx (not .xls) and try again."
        ) from e

    if spec.sheet_name not in wb.sheetnames:
        raise ConversionError(
            f"Workbook is missing required sheet {spec.sheet_name!r}. "
            f"Found sheets: {wb.sheetnames}. Re-download the template."
        )
    ws = wb[spec.sheet_name]

    # Row 1: header validation. In read_only mode openpyxl returns
    # EmptyCell objects for blanks, which don't carry .column — so we
    # enumerate the row index ourselves rather than rely on the cell's
    # column attribute.
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    uploaded_columns = [
        (cell.value, col_idx + 1)
        for col_idx, cell in enumerate(header_row)
        if cell.value is not None
    ]
    uploaded_names = {str(c[0]).strip() for c in uploaded_columns}

    expected_required = {c.name for c in spec.columns if c.required}
    missing_required = expected_required - uploaded_names
    if missing_required:
        raise ConversionError(
            f"Required column(s) missing from row 1 of sheet {spec.sheet_name!r}: "
            f"{sorted(missing_required)}. Re-download the template."
        )
    # Tolerate extra columns silently — schema evolution. Just won't write them.

    # Map column name → ColumnSpec, and column name → Excel column index
    col_specs_by_name: dict[str, ColumnSpec] = {c.name: c for c in spec.columns}
    col_idx_by_name: dict[str, int] = {
        str(name).strip(): idx for name, idx in uploaded_columns
        if str(name).strip() in col_specs_by_name
    }

    # Build per-column lists. Skip header (row 1), hint (row 2), example (row 3).
    columns_data: dict[str, list] = {name: [] for name in col_idx_by_name}
    skipped_cells = 0
    row_count = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=4), start=4):
        # Snapshot the cells we care about. EmptyCells from openpyxl
        # read-only mode don't carry .column, so enumerate the row's
        # cells positionally — same index basis as header_row above.
        values_by_idx: dict[int, object] = {
            col_idx + 1: cell.value for col_idx, cell in enumerate(row)
        }
        # Skip fully-blank rows
        if all(values_by_idx.get(idx) in (None, "") for idx in col_idx_by_name.values()):
            continue

        for name, col_idx in col_idx_by_name.items():
            col_spec = col_specs_by_name[name]
            raw = values_by_idx.get(col_idx)
            coerced = _coerce_cell(raw, col_spec, row_num)
            if coerced is None:
                if col_spec.required:
                    raise ConversionError(
                        f"Row {row_num}: required column {name!r} is empty."
                    )
                skipped_cells += 1
            columns_data[name].append(coerced)

        row_count += 1

    if row_count == 0:
        raise ConversionError(
            f"Sheet {spec.sheet_name!r} contains no data rows. "
            f"Add at least one row of data starting at row 4."
        )

    # Build the PyArrow table — preserves column order from the template
    arrays = []
    field_specs = []
    for col in spec.columns:
        if col.name not in columns_data:
            continue
        pa_type = pyarrow_type_for(col.logical_type)
        arrays.append(pa.array(columns_data[col.name], type=pa_type))
        # nullable=True except for required columns (which we already
        # rejected on empty; this is a defense-in-depth signal in the
        # parquet schema)
        field_specs.append(pa.field(col.name, pa_type, nullable=not col.required))

    table = pa.Table.from_arrays(arrays, schema=pa.schema(field_specs))

    out_path = parquet_path_for(record.upload_id, record.bronze_table)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    pq.write_table(table, out_path)

    return ConversionResult(
        parquet_path=out_path,
        row_count=row_count,
        skipped_cells=skipped_cells,
    )

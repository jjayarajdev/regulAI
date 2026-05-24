"""Phase 2A upload tests — no live Dagster, no live Snowflake.

Covers:
  - Schema registry sanity
  - Template generation (openpyxl Workbook bytes parse back cleanly)
  - Storage IO (new id format, registry round-trip, status updates)
  - Excel → Parquet conversion (happy path, schema mismatch, type errors)
  - FastAPI routes (upload, list, process — Dagster mocked offline)

A separate end-to-end test that spins up `dagster dev` + a live
Snowflake stage is out of scope here; that's the manual smoke test
documented in docs/dagster-orchestration.md.
"""

import datetime as dt
import io
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
import pyarrow.parquet as pq


# ── Schema registry ────────────────────────────────────────────────


def test_phase2a_registers_at_least_one_template():
    """The whole upload flow is gated on the registry. If TEMPLATES is
    empty the UI dropdown is empty too."""
    from packages.uploads import TEMPLATES

    assert len(TEMPLATES) >= 1
    # All templates have at least one required column — otherwise the
    # validation gate has nothing to enforce.
    for t in TEMPLATES:
        required = [c for c in t.columns if c.required]
        assert required, f"Template {t.bronze_table} has no required columns"


def test_get_template_returns_none_for_unknown():
    from packages.uploads.schemas import get_template

    assert get_template("does_not_exist") is None


def test_template_columns_have_valid_logical_types():
    """Catches typos in schemas.py: 'strng' instead of 'string', etc."""
    from packages.uploads.schemas import TEMPLATES, pyarrow_type_for

    for t in TEMPLATES:
        for c in t.columns:
            # If logical_type is bogus, pyarrow_type_for will KeyError
            pyarrow_type_for(c.logical_type)


# ── Template generation ────────────────────────────────────────────


def test_generate_template_workbook_parses_back():
    """The bytes we serve to the user must open as a valid .xlsx with
    the expected sheet structure."""
    from packages.uploads import generate_template_workbook

    xlsx_bytes = generate_template_workbook("gw_pc_policy")
    wb = openpyxl.load_workbook(filename=io.BytesIO(xlsx_bytes))
    assert "pc_policy" in wb.sheetnames
    assert "_README" in wb.sheetnames
    # Row 1 contains the column headers
    ws = wb["pc_policy"]
    header_values = [c.value for c in ws[1]]
    assert "id" in header_values
    assert "policynumber" in header_values


def test_generate_template_unknown_table_raises():
    from packages.uploads import generate_template_workbook

    with pytest.raises(KeyError):
        generate_template_workbook("not_a_table")


# ── Storage IO ─────────────────────────────────────────────────────


@pytest.fixture
def upload_storage_tmp(tmp_path, monkeypatch):
    """Redirect upload storage to a tmp dir so tests don't trample
    real uploads under materialized/uploads/."""
    import packages.uploads.storage as storage

    monkeypatch.setattr(storage, "UPLOADS_DIR", tmp_path)
    monkeypatch.setattr(storage, "_REGISTRY_PATH", tmp_path / "_registry.json")
    return tmp_path


def test_new_upload_id_is_sortable_and_unique(upload_storage_tmp):
    from packages.uploads.storage import new_upload_id

    ids = {new_upload_id() for _ in range(20)}
    assert len(ids) == 20, "expected 20 unique ids; got collisions"
    for i in ids:
        assert i.startswith("upl_")


def test_registry_round_trip(upload_storage_tmp):
    from packages.uploads.storage import (
        UploadRecord,
        list_uploads,
        new_upload_dir,
        new_upload_id,
        record_upload,
        update_upload_status,
    )

    upload_id = new_upload_id()
    new_upload_dir(upload_id)
    rec = UploadRecord(
        upload_id=upload_id, filename="x.xlsx",
        bronze_table="gw_pc_policy", bytes_size=4242,
    )
    record_upload(rec)
    listed = list_uploads()
    assert len(listed) == 1
    assert listed[0].upload_id == upload_id
    assert listed[0].status == "uploaded"

    update_upload_status(upload_id, status="converted", row_count=7)
    listed = list_uploads()
    assert listed[0].status == "converted"
    assert listed[0].row_count == 7


def test_registry_recovers_from_corruption(upload_storage_tmp):
    """A malformed _registry.json mustn't crash list_uploads()."""
    (upload_storage_tmp / "_registry.json").write_text("not valid json", encoding="utf-8")
    from packages.uploads.storage import list_uploads

    assert list_uploads() == []


# ── Excel → Parquet conversion ─────────────────────────────────────


def _make_test_upload(upload_storage_tmp, extra_data_rows: list[list] | None = None):
    """Generate a template, optionally append data rows, return UploadRecord."""
    from packages.uploads import (
        UploadRecord,
        generate_template_workbook,
        new_upload_dir,
        new_upload_id,
        record_upload,
    )

    xlsx = generate_template_workbook("gw_pc_policy")
    upload_id = new_upload_id()
    root = new_upload_dir(upload_id)
    xlsx_path = root / "original" / "gw_pc_policy_template.xlsx"
    xlsx_path.write_bytes(xlsx)

    if extra_data_rows:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["pc_policy"]
        for row in extra_data_rows:
            ws.append(row)
        wb.save(xlsx_path)

    rec = UploadRecord(
        upload_id=upload_id, filename="gw_pc_policy_template.xlsx",
        bronze_table="gw_pc_policy", bytes_size=xlsx_path.stat().st_size,
    )
    record_upload(rec)
    return rec


def test_conversion_happy_path(upload_storage_tmp):
    from packages.uploads import convert_uploaded_xlsx_to_parquet

    rec = _make_test_upload(
        upload_storage_tmp,
        extra_data_rows=[
            [2000999, "pc:test", 300999, 500999, "HO-T001",
             "2025-06-01", "2025-06-01", "2025-06-01 09:00:00", "2025-06-01 09:00:00", 0],
            [2001000, None, 300999, None, "HO-T002",
             "2025-06-15", None, None, None, None],
        ],
    )
    result = convert_uploaded_xlsx_to_parquet(rec)
    assert result.row_count == 2

    table = pq.read_table(result.parquet_path)
    assert table.column_names == [c.name for c in __import__("packages.uploads.schemas", fromlist=["get_template"]).get_template("gw_pc_policy").columns]
    # Required columns aren't null on either row
    ids = table["id"].to_pylist()
    assert ids == [2000999, 2001000]
    # Optional columns are None where left blank
    assert table["producercode_id"].to_pylist() == [500999, None]


def test_conversion_rejects_missing_required_column(upload_storage_tmp):
    """If the user nukes a required column from the template, the
    conversion fails with a clear message before anything writes."""
    from packages.uploads import (
        UploadRecord,
        convert_uploaded_xlsx_to_parquet,
        generate_template_workbook,
        new_upload_dir,
        new_upload_id,
        record_upload,
    )
    from packages.uploads.xlsx_to_parquet import ConversionError

    upload_id = new_upload_id()
    root = new_upload_dir(upload_id)
    xlsx_path = root / "original" / "broken.xlsx"
    # Write a template, then DELETE the 'policynumber' header
    xlsx_path.write_bytes(generate_template_workbook("gw_pc_policy"))
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["pc_policy"]
    for cell in ws[1]:
        if cell.value == "policynumber":
            cell.value = None
    ws.append([2000999, None, 300999, None,
               # policynumber col now blank — header missing
               None, "2025-06-01", None, None, None, 0])
    wb.save(xlsx_path)

    rec = UploadRecord(
        upload_id=upload_id, filename="broken.xlsx",
        bronze_table="gw_pc_policy", bytes_size=xlsx_path.stat().st_size,
    )
    record_upload(rec)

    with pytest.raises(ConversionError, match="missing"):
        convert_uploaded_xlsx_to_parquet(rec)


def test_conversion_rejects_empty_required_cell(upload_storage_tmp):
    """A required cell left blank in a data row gets a clear row-level error."""
    from packages.uploads import convert_uploaded_xlsx_to_parquet
    from packages.uploads.xlsx_to_parquet import ConversionError

    rec = _make_test_upload(
        upload_storage_tmp,
        extra_data_rows=[
            [2000999, None, 300999, None, None,  # policynumber required — None breaks
             "2025-06-01", None, None, None, 0],
        ],
    )
    with pytest.raises(ConversionError, match="required column .*policynumber.* is empty"):
        convert_uploaded_xlsx_to_parquet(rec)


def test_conversion_rejects_bad_integer(upload_storage_tmp):
    from packages.uploads import convert_uploaded_xlsx_to_parquet
    from packages.uploads.xlsx_to_parquet import ConversionError

    rec = _make_test_upload(
        upload_storage_tmp,
        extra_data_rows=[
            ["NOT_AN_INT", None, 300999, None, "HO-T001",
             "2025-06-01", None, None, None, 0],
        ],
    )
    with pytest.raises(ConversionError, match="cannot parse.*as integer"):
        convert_uploaded_xlsx_to_parquet(rec)


def test_conversion_skips_blank_trailing_rows(upload_storage_tmp):
    from packages.uploads import convert_uploaded_xlsx_to_parquet

    rec = _make_test_upload(
        upload_storage_tmp,
        extra_data_rows=[
            [2000999, None, 300999, None, "HO-T001",
             "2025-06-01", None, None, None, 0],
            [None] * 10,  # totally blank row — should be skipped, not error
        ],
    )
    result = convert_uploaded_xlsx_to_parquet(rec)
    assert result.row_count == 1


# ── FastAPI routes ─────────────────────────────────────────────────


@pytest.fixture
def client(upload_storage_tmp):
    from fastapi.testclient import TestClient

    from api.main import app
    return TestClient(app)


def test_get_upload_templates_list(client):
    r = client.get("/api/admin/upload-templates")
    assert r.status_code == 200
    body = r.json()
    assert body["templates"]
    names = {t["bronze_table"] for t in body["templates"]}
    assert "gw_pc_policy" in names


def test_download_template_serves_xlsx(client):
    r = client.get("/api/admin/upload-templates/gw_pc_policy")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    # Parse the bytes — must be a valid workbook
    wb = openpyxl.load_workbook(filename=io.BytesIO(r.content))
    assert "pc_policy" in wb.sheetnames


def test_download_unknown_template_404(client):
    r = client.get("/api/admin/upload-templates/does_not_exist")
    assert r.status_code == 404


def test_admin_upload_page_renders(client):
    r = client.get("/admin/upload")
    assert r.status_code == 200
    assert "Upload data" in r.text
    assert "Download" in r.text


def test_upload_endpoint_rejects_csv(client):
    r = client.post(
        "/api/admin/uploads",
        data={"bronze_table": "gw_pc_policy"},
        files={"file": ("data.csv", b"a,b,c\n1,2,3", "text/csv")},
    )
    assert r.status_code == 400
    assert ".xlsx" in r.json()["detail"]


def test_upload_endpoint_rejects_unknown_table(client):
    from packages.uploads import generate_template_workbook

    r = client.post(
        "/api/admin/uploads",
        data={"bronze_table": "does_not_exist"},
        files={"file": ("data.xlsx", generate_template_workbook("gw_pc_policy"), "application/xlsx")},
    )
    assert r.status_code == 400
    assert "Unknown bronze_table" in r.json()["detail"]


def test_upload_endpoint_happy_path(client):
    """End-to-end: upload an xlsx with data, get back a converted record."""
    from packages.uploads import generate_template_workbook

    # Build an in-memory xlsx with one data row
    xlsx = generate_template_workbook("gw_pc_policy")
    wb = openpyxl.load_workbook(filename=io.BytesIO(xlsx))
    ws = wb["pc_policy"]
    ws.append([2000999, "pc:test", 300999, 500999, "HO-API-001",
               "2025-06-01", None, None, None, 0])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        "/api/admin/uploads",
        data={"bronze_table": "gw_pc_policy"},
        files={"file": ("via_api.xlsx", buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    assert body["row_count"] == 1
    assert body["upload"]["status"] == "converted"
    assert body["upload"]["bronze_table"] == "gw_pc_policy"


def test_list_uploads_after_upload(client):
    """After a successful upload, the history endpoint includes it."""
    from packages.uploads import generate_template_workbook

    xlsx = generate_template_workbook("gw_pc_policy")
    wb = openpyxl.load_workbook(filename=io.BytesIO(xlsx))
    ws = wb["pc_policy"]
    ws.append([2000999, None, 300999, None, "HO-L-001",
               "2025-06-01", None, None, None, 0])
    buf = io.BytesIO()
    wb.save(buf)

    client.post(
        "/api/admin/uploads",
        data={"bronze_table": "gw_pc_policy"},
        files={"file": ("h.xlsx", buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    r = client.get("/api/admin/uploads")
    assert r.status_code == 200
    body = r.json()
    assert body["uploads"]
    assert body["uploads"][0]["filename"] == "h.xlsx"


def test_process_upload_returns_503_when_dagster_offline(client):
    """The Process button must surface a clean 503 when Dagster isn't
    running — not a 500 stack trace."""
    from packages.uploads import (
        UploadRecord,
        new_upload_dir,
        new_upload_id,
        record_upload,
        update_upload_status,
    )
    from packages.scheduling.dagster_client import DagsterUnreachable

    upload_id = new_upload_id()
    new_upload_dir(upload_id)
    rec = UploadRecord(
        upload_id=upload_id, filename="x.xlsx",
        bronze_table="gw_pc_policy", bytes_size=10,
    )
    record_upload(rec)
    update_upload_status(upload_id, status="converted", row_count=1)

    with patch(
        "packages.scheduling.dagster_client.launch_run",
        side_effect=DagsterUnreachable("connection refused"),
    ):
        r = client.post(f"/api/admin/uploads/{upload_id}/process")
    assert r.status_code == 503
    assert "Dagster" in r.json()["detail"]


def test_process_upload_404_for_missing_upload(client):
    r = client.post("/api/admin/uploads/upl_99991231_deadbeef/process")
    assert r.status_code == 404
